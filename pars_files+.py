import os, os.path, sys
import time
import openpyxl
from win32com.client import Dispatch
sys.setrecursionlimit(100000)

shell = Dispatch("Shell.Application")

dest_path = input('Введите стартовую директорию: ')

count = 2


def find_owner(filename):
    return os.stat(filename).st_uid


def all_paths(dest_path, count=2):
    ns = shell.NameSpace(os.path.normpath(dest_path))
    for elem in os.listdir(dest_path):
        dict = {}
        elem = os.path.join(dest_path, elem)
        if os.path.isfile(elem):
            elem_end = elem.split('.')
            elem_end = elem_end[1]
            elem_name = elem.split('\\')
            elem_name = elem_name[-1]
            for i in ns.Items():
                if str(i) == elem_name:
                    for j in range(0, 49):
                        dict[ns.GetDetailsOf(j, j)] = ns.GetDetailsOf(i, j)
            t1 = time.ctime(os.path.getmtime(elem))
            t2 = time.ctime(os.path.getctime(elem))
            elem_autor = dict.get('Авторы')
            sheet.cell(row=count, column=1).value = elem_name
            sheet.cell(row=count, column=2).value = os.path.relpath(elem)
            sheet.cell(row=count, column=3).value = elem_end
            sheet.cell(row=count, column=4).value = elem_autor
            sheet.cell(row=count, column=5).value = t1
            sheet.cell(row=count, column=6).value = t2
            sheet.cell(row=count, column=7).value = os.path.getsize(elem)
            sheet.cell(row=count, column=8).value = os.path.getsize(elem) / 1024
            sheet.cell(row=count, column=9).value = os.path.getsize(elem) / 1048576
            count += 1
        if os.path.isdir(elem):
            all_paths(elem)


excel_doc = openpyxl.load_workbook('nrapars.xlsx')
sheet = excel_doc['1'] #Здесь поменять имя листа. Но он уже должен существовать!

all_paths(dest_path)

excel_doc.save('nrapars.xlsx')
